"""
step4_annotate_policies.py
==========================
Updated annotation script that uses the trained GPU classifier
instead of Mistral for sentence labelling.

If the classifier service is not running, it falls back to
Mistral automatically (Phase 1 / bootstrap mode).

Usage:
    # With GPU classifier running on port 8002:
    python step4_annotate_policies.py --input ./exported_files --limit 10

    # Force Mistral fallback (Phase 1 bootstrap):
    python step4_annotate_policies.py --input ./exported_files --limit 10 --mistral-only

Requirements:
    pip install pypdf "pdfplumber==0.10.4" openpyxl requests

The classifier service (step3_classifier_api.py) must be running on
your GPU server at CLASSIFIER_URL before calling this.
"""

import re
import logging
import sys
import json
import argparse
import traceback
import requests
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Tuple, Optional

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from pypdf import PdfReader
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

try:
    import fitz  # pymupdf
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    import spacy as _spacy
    _nlp = _spacy.load("en_core_web_sm")
    HAS_SPACY = True
except Exception:
    HAS_SPACY = False

try:
    import networkx as _nx
    HAS_NX = True
except ImportError:
    HAS_NX = False

if not HAS_PDFPLUMBER and not HAS_PYPDF and not HAS_PYMUPDF:
    print("ERROR: pip install pymupdf")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# ── Service URLs ─────────────────────────────────────────────────────────────
CLASSIFIER_URL = "http://140.203.155.230:8002"   # your GPU server
MISTRAL_URL    = "http://localhost:11434/api/chat"

LABEL_COLOURS = {
    "Level: Policy Action":  "D9EAD3",
    "Level: Policy Outcome": "CFE2F3",
    "Level: Unsure":         "FFF2CC",
    "Class: Area":           "FCE5CD",
    "Class: Emissions":      "D0E4C8",
    "Class: Site Status":    "EAD1DC",
    "Class: Spending":       "C9DAF8",
    "Class: Policy Action":  "D9D2E9",
    "Class: Knowledge Resource":  "E6F3FF",
    "Class: Practical Resource":  "FDE8D8",
    "Class: Environment Quality": "E2F0D9",
    "Class: Miscellaneous":  "E8E8E8",
}

# ── Indirect sentence signals (rule-based, same as classifier API) ───────────
METRIC_PATTERNS = [
    # 1. Quantified values — handles EN/FR/DE/SK/ES number formats
    # Area units: ha, hectares, hektárov (SK), hectáreas (ES), Hektar (DE)
    r"\b\d[\d,\.\s]*\s*(ha|hectares?|hektár(ov)?|hectáreas?|Hektar)\b",
    r"\b\d[\d,\.]*\s*(km2?|acres?)\b",
    r"\b\d[\d,\.]*\s*(%|percent|tonne|MtCO2|t\s*CO2|million\s*tonnes?)\b",
    # Currency: € before or after number (EN: €1,500 / FR: 1 500 €)
    r"€\s*\d[\d,\.\s]*",
    r"\b\d[\d,\.\s]*\s*€",
    r"\b\d[\d,\.]*\s*(million|billion|thousand)\s*(euro|euros?|tonnes?|ha)\b",
    r"\bby\s+(20\d{2}|19\d{2})\b",
    r"\bover\s+\d[\d,\.]*\s+years?\b",
    r"\b\d[\d,\.]*\s*(sites?|bogs?|SACs?|NHAs?|SPAs?)\b",
    r"\b20[12]\d\s+to\s+20[12]\d\b",
    r"\bper\s+annum\b",
    # 2. Named deliverables
    r"\b(publish|produce|prepare|develop|establish|create|introduce|launch|implement)\b.{0,60}\b(plan|report|database|survey|scheme|programme|strategy|guidance|review|framework|register|code)\b",
    r"\b(national|pilot|annual|interim)\s+(plan|report|database|survey|scheme|programme|strategy|review)\b",
    # 3. Named responsible body making a commitment
    r"\b(DAFM|NPWS|EPA|OPW|Coillte|Bord\s+na\s+M.na|DAHG|IPCC|the\s+Minister|the\s+Department|the\s+Government|the\s+State|local\s+authorit)\b.{0,80}\b(shall|will|must|commit|undertake|ensure|provide|deliver|carry\s+out|establish|implement)\b",
]
_metric_re = re.compile("|".join(METRIC_PATTERNS), re.IGNORECASE)

INDIRECT_PATTERNS = [
    r"regulation\s*\(eu\)\s*\d{4}/\d+",
    r"regulation\s*\(ec\)\s*\d{4}/\d+",
    r"directive\s*\d{4}/\d+",
    r"\barticle\s+\d+\s+of\s+(regulation|directive|the)",
    r"pursuant\s+to\s+(regulation|directive|article|the\s+act)",
    r"in\s+accordance\s+with\s+(regulation|directive|article)",
    r"as\s+required\s+by\s+(regulation|directive|the\s+act)",
    r"under\s+regulation\s*\(",
    r"member\s+states?\s+shall",
]
_indirect_re = re.compile("|".join(INDIRECT_PATTERNS), re.IGNORECASE)


# ── Hedge word detection (auto-assign Level: Unsure) ─────────────────────────
HEDGE_PATTERNS = [
    # Uncertain modal verbs
    r"\bcould\b",
    r"\bmay\b",
    r"\bpotentially\b",
    r"\bsubject\s+to\b",
    r"\bif\s+funding\b",
    r"\bwhere\s+possible\b",
    r"\bwhere\s+appropriate\b",
    r"\bintend\s+to\b",
    r"\baim\s+to\b",
    # Deferral language — sounds like commitment but is not verifiable
    r"\bwill\s+be\s+considered\b",
    r"\bconsideration\s+will\s+be\s+given\b",
    r"\bwill\s+be\s+explored\b",
    r"\bwill\s+be\s+examined\b",
    r"\bshould\s+be\s+considered\b",
    r"\bif\s+appropriate\b",
    r"\bif\s+feasible\b",
    # Background/descriptive language
    r"\bit\s+is\s+noted\s+that\b",
    r"\bexperience\s+has\s+shown\b",
    r"\bcan\s+(\w+\s+)?(help|provide|contribute|support|assist|offer|play)\b",
    # "To date", "As of [year]" — historical snapshots, not commitments
    r"\bto\s+date[,.]?\b",
    r"\bas\s+of\s+(late\s+|early\s+|mid\s+)?(19|20)\d{2}\b",
    # "it is estimated / it is recognised / it is acknowledged"
    r"\bit\s+is\s+(estimated|recognised|acknowledged|understood|expected)\b",
]
_hedge_re = re.compile("|".join(HEDGE_PATTERNS), re.IGNORECASE)

BOILERPLATE_RE = re.compile(
    r"^(table of contents?|list of (tables?|figures?)|"
    r"page \d+|figure \d+|annex [a-z\d]|"
    r"copyright|all rights reserved|isbn|doi:)",
    re.IGNORECASE,
)

# Footnote/bibliography references — not policy sentences
FOOTNOTE_RE = re.compile(
    r"^\d{1,2}\s+[A-Z][a-z]+,\s*[A-Z]"
    r"|^\d{1,2}\s+\d\s+[A-Z]"
    r"|^\d{1,2}\s+(Figures|CRF|Table\s+\d|Source|See\s+also|Based\s+on)\b"
    r"|\bop\.\s*cit\b|\bibid\b"
    r"|^\[\d+\]"
    r"|\bISBN\b|\bDOI\b|\bISSN\b",
    re.IGNORECASE,
)

# Page headers — repeated at top/bottom of each page
PAGE_HEADER_RE = re.compile(
    r"^NATIONAL\s+PEATLANDS\s+STRATEGY\s+[/\d]"
    r"|^\d{3}/\d{3}\b"
    r"|^PART\s+\d+\s*$"
    r"|^APPENDIX\s+[IVX\d]+\s*$"
    # UNCCD/UNFCCC document page headers e.g. "ICCD/COP(13)/21/Add.1"
    r"|^ICCD/"
    r"|^FCCC/"
    r"|^UNFCCC/"
    r"|^CBD/",
    re.IGNORECASE,
)

# Contact details / cover page info
CONTACT_RE = re.compile(
    r"t:\s*\+\d{1,3}"
    r"|e:\s*\w+@\w+"
    r"|w:\s*www\."
    r"|\+353"
    r"|@[a-z]+\.gov\.ie"
    r"|@[a-z]+\.ie\b",
    re.IGNORECASE,
)

# Table of contents patterns
TOC_RE = re.compile(
    r"(\b\d{3}\b.*){3,}"
    r"|\bCONTENTS\b.*\bPART\b"
    r"|\b\d+\.\d+\s+[A-Z].*\b\d{3}\b.*\b\d+\.\d+",
    re.IGNORECASE,
)

# Table cells read as prose (ecosystem services table etc)
TABLE_CONTENT_RE = re.compile(
    r"(Cultural tradition|ecosystem services|Clean water|Climate regulation"
    r"|Flood and erosion|Wildfire hazard|Nutrient cycling|Carbon sequestration"
    r"|Biodiversity|Contaminant removal|Water filtration).{0,30}"
    r"(Cultural tradition|ecosystem services|Clean water|Climate regulation"
    r"|Flood and erosion|Wildfire hazard|Nutrient cycling|Carbon sequestration"
    r"|Biodiversity|Contaminant removal|Water filtration)",
    re.IGNORECASE,
)

# Document title pages
TITLE_PAGE_RE = re.compile(
    r"^MANAGING\s+IRELAND"
    r"|^A\s+National\s+Peatlands\s+Strategy\s+\d{4}$"
    r"|^National\s+Peatlands\s+Strategy\s+\d{4}$",
    re.IGNORECASE,
)

# Photo credits / acknowledgements lines
PHOTO_CREDITS_RE = re.compile(
    r"^(NPWS|Bord\s+na\s+M.na|Coillte|RPS)[,\s]"
    r"|photographs?\s+courtesy\s+of"
    r"|photos?\s+by\b"
    r"|©\s*\d{4}",
    re.IGNORECASE,
)

# Footnote definitions starting with *
FOOTNOTE_DEF_RE = re.compile(
    r'^[*]\s*(non-sustainable|sustainable|means\s+that|refers\s+to)',
    re.IGNORECASE,
)

# Background science / definitional sentences — not policy commitments
BACKGROUND_SCIENCE_RE = re.compile(
    r"^Peatlands?\s+(have\s+been\s+in\s+the\s+Irish\s+landscape"
    r"|are\s+wetland\s+ecosystems"
    r"|are\s+the\s+country.s\s+last\s+great"
    r"|form\s+our\s+oldest"
    r"|are\s+principally\s+bogs"
    r"|only\s+fed\s+by\s+precipitation"
    r"|play\s+an\s+important\s+part\s+in\s+maintaining)"
    r"|^As\s+they\s+develop,\s+peatlands"
    r"|^Over\s+a\s+long\s+period\s+of\s+time,\s+peatlands"
    r"|^Bogs?\s+are\s+peatlands\s+only\s+fed"
    r"|^Fens?\s+are\s+peatlands\s+that\s+in\s+addition"
    r"|^A\s+raised\s+bog\s+is\s+a\s+bog\s+shaped"
    r"|^A\s+blanket\s+bog\s+is\s+a\s+bog\s+that\s+covers"
    r"|^Peat\s+(is\s+used\s+in\s+horticulture|soils?\s+cover\s+around)",
    re.IGNORECASE,
)

# Past-tense historical fact patterns — these are context, not commitments
HISTORICAL_RE = re.compile(
    r"\b(was|were|had|has\s+been|have\s+been)\b.{0,40}"
    r"\b(established|designated|nominated|created|founded|introduced|"
    r"published|adopted|signed|ratified|formed|set\s+up)\b"
    r"|\bbetween\s+\d{4}\s+and\s+\d{4}\b"
    r"|\bin\s+(the\s+)?(19|20)\d{2}\b.{0,30}\b(was|were|had)\b"
    r"|\bthe\s+original\s+(extent|area|coverage)\b"
    r"|\bit\s+(is|was)\s+estimated\s+that\b"
    r"|\bas\s+of\s+(late\s+|early\s+|mid\s+)?(19|20)\d{2}\b"

    r"|\bhistorically[,.]\b"
    r"|\bin\s+recent\s+(years|decades|times)[,.]\b",
    re.IGNORECASE,
)


# ── PDF extraction ────────────────────────────────────────────────────────────
def _clean_text(text: str) -> str:
    # Keep ASCII + common Unicode (accented, smart quotes, dashes)
    text = re.sub(r'[^\x09\x0A\x20-\x7E\xA0-\u024F\u2013\u2014\u2018-\u201D\u2026]', ' ', text)
    text = re.sub(r'[ \t]{2,}', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def clean_chunk_text(text: str) -> str:
    """
    Minimal safe cleaning for PDF extracted text.
    Works for both single-column and two-column layouts.
    Does NOT join lines aggressively — that destroys single-column text.
    Fixes: hyphenated word splits, apostrophe-s splits,
    single-newline line wraps, trailing garbage, section headers.
    """
    # Fix hyphenated word splits across lines (e.g. "afforesta-\ntion")
    text = re.sub(r"-\s*\n\s*([a-z])", r"\1", text)

    # Fix apostrophe-s split (e.g. "biodiversity\ns " → "biodiversity\'s ")
    text = re.sub(r"\n(s[\s])", r"\'\1", text)

    # Replace single newlines with space (PDF line wraps within paragraph)
    # Preserve double newlines (paragraph breaks)
    text = re.sub(r"(?<!\n)\n(?!\n)", " ", text)

    # Strip leading section headers e.g. "FORESTRY - ACTION A7 "
    text = re.sub(
        r"^[A-Z][A-Z\s/\-&]+\s+(ACTION|PRINCIPLE|POLICY|ACTIONS|SECTION)\s+[A-Z]?\d+\s+",
        "", text
    ).strip()

    # Truncate trailing garbage after last sentence end (max 25 chars)
    last_end = max(text.rfind("."), text.rfind("!"), text.rfind("?"))
    if last_end > len(text) * 0.5:
        after = text[last_end + 1:].strip()
        if len(after) <= 25:
            text = text[:last_end + 1]

    # Normalise multiple spaces
    text = re.sub(r"  +", " ", text).strip()
    return text



def _is_garbage(text: str) -> bool:
    if not text or len(text.strip()) < 20:
        return True
    # Must have at least 3 real words
    return len(re.findall(r'[a-zA-Z]{3,}', text)) < 3


def _extract_page_pymupdf(page) -> str:
    """
    Robust pymupdf text extraction.
    Tries multiple methods to handle unusual PDF encodings:
    - Method 1: blocks (best reading order, works for most PDFs)
    - Method 2: rawdict (works for Type3 fonts, unusual encodings)
    - Method 3: plain text (last resort)
    Handles both single-column and multi-column layouts correctly.
    """
    # Method 1: blocks — best reading order, handles columns automatically
    try:
        blocks = page.get_text("blocks")
        text_blocks = [b for b in blocks if b[6] == 0 and b[4].strip()]
        if text_blocks:
            text_blocks.sort(key=lambda b: (round(b[1] / 20) * 20, b[0]))
            text = "\n\n".join(b[4].strip() for b in text_blocks)
            if len(re.findall(r'[a-zA-Z]{3,}', text)) >= 3:
                return text
    except Exception:
        pass

    # Method 2: rawdict — handles Type3 fonts and non-standard encodings
    try:
        rawdict = page.get_text("rawdict")
        lines = []
        for block in sorted(
            rawdict.get("blocks", []),
            key=lambda b: (round(b.get("bbox", [0,0,0,0])[1] / 20) * 20,
                           b.get("bbox", [0,0,0,0])[0])
        ):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                line_text = "".join(
                    span.get("text", "") for span in line.get("spans", [])
                )
                if line_text.strip():
                    lines.append(line_text.strip())
        if lines:
            text = "\n".join(lines)
            if len(re.findall(r'[a-zA-Z]{3,}', text)) >= 3:
                return text
    except Exception:
        pass

    # Method 3: plain text fallback
    try:
        text = page.get_text("text")
        if text.strip():
            return text
    except Exception:
        pass

    return ""


def extract_text_from_pdf(pdf_path: Path) -> List[Dict]:
    """
    Extract text page by page from a PDF using pymupdf.
    Handles single-column, multi-column, Type3 fonts, and
    unusual PDF encodings via three internal fallback methods.
    """
    pages = []
    try:
        doc = fitz.open(str(pdf_path))
        for i in range(len(doc)):
            page = doc[i]
            try:
                text = _clean_text(_extract_page_pymupdf(page))
                source = "pymupdf"
            except Exception:
                text, source = "", "skipped"
            pages.append({
                "page_num": i + 1,
                "text":     "" if _is_garbage(text) else text,
                "source":   "skipped_graphic" if _is_garbage(text) else source,
            })
        doc.close()

        # Warn if all pages were empty (truly scanned/image-only PDF)
        extracted = [p for p in pages if p["source"] != "skipped_graphic"]
        if not extracted:
            logging.warning(
                f"{pdf_path.name}: no text extracted on any page. "
                f"PDF may be scanned — consider OCR preprocessing."
            )
    except Exception as e:
        logging.error(f"Failed to open {pdf_path.name}: {e}")

    return pages

def _split_sentences(text: str) -> list:
    """
    Split text into complete sentences.
    Never cuts mid-sentence — always splits at . ! ? ; or paragraph breaks.
    """
    paragraphs = re.split(r'\n{2,}', text)
    sentences = []
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        # Split on sentence-ending punctuation followed by space + capital/digit/bracket
        # or on semicolons separating policy sub-clauses
        parts = re.split(
            r'(?<=[.!?])\s+(?=[A-Z(\[\d])|(?<=;)\s+(?=[A-Z(\[A-Za-z])',
            para
        )
        for part in parts:
            part = part.strip()
            if part:
                sentences.append(part)
    return sentences


def chunk_pages(pages: List[Dict],
                chunk_size: int = 500,
                chunk_overlap: int = 50) -> List[Dict]:
    """
    Chunk pages into complete-sentence chunks.
    NEVER splits mid-sentence — only splits at sentence boundaries.
    If a single sentence exceeds chunk_size it is kept whole (not cut).
    Multiple short sentences are merged until they approach chunk_size.
    """
    chunks, idx = [], 0
    for page in pages:
        text = page["text"].strip()
        if not text:
            continue

        sentences = _split_sentences(text)
        current = ""

        for sent in sentences:
            sent = sent.strip()
            if not sent:
                continue
            if not current:
                current = sent
            elif len(current) + 1 + len(sent) <= chunk_size:
                current = current + " " + sent
            else:
                # Save current chunk — always complete sentences
                if len(current) >= 30:
                    chunks.append({
                        "text":        clean_chunk_text(current),
                        "page":        page["page_num"],
                        "chunk_index": idx,
                        "source":      page["source"],
                    })
                    idx += 1
                # Start new chunk — keep sentence whole even if > chunk_size
                current = sent

        if current.strip() and len(current.strip()) >= 30:
            chunks.append({
                "text":        clean_chunk_text(current.strip()),
                "page":        page["page_num"],
                "chunk_index": idx,
                "source":      page["source"],
            })
            idx += 1

    return chunks



# ── Pre-filter (rule-based, no model needed) ──────────────────────────────────
def pre_filter(chunks: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """
    Split chunks into (to_classify, skip).
    Skip: boilerplate, too short, pure numbers, historical facts.
    """
    to_classify, skip = [], []
    for c in chunks:
        text = c["text"]
        if len(text) < 40:
            skip.append({**c, "skip_reason": "too_short"})
        elif BOILERPLATE_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "boilerplate"})
        elif FOOTNOTE_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "footnote"})
        elif PAGE_HEADER_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "page_header"})
        elif CONTACT_RE.search(text.strip()):
            skip.append({**c, "skip_reason": "contact_info"})
        elif TOC_RE.search(text.strip()):
            skip.append({**c, "skip_reason": "table_of_contents"})
        elif TABLE_CONTENT_RE.search(text.strip()):
            skip.append({**c, "skip_reason": "table_content"})
        elif TITLE_PAGE_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "title_page"})
        elif PHOTO_CREDITS_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "photo_credits"})
        elif FOOTNOTE_DEF_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "footnote_definition"})
        elif BACKGROUND_SCIENCE_RE.match(text.strip()):
            skip.append({**c, "skip_reason": "background_science"})
        elif re.match(r'^[\d\s.,;:()-]+$', text.strip()):
            skip.append({**c, "skip_reason": "numeric_only"})
        elif HISTORICAL_RE.search(text):
            skip.append({**c, "skip_reason": "historical_context"})
        else:
            to_classify.append(c)
    return to_classify, skip


# ── PageRank dependency scoring (SAPC Optimization 2) ────────────────────────
def _compute_pagerank_scores(texts: list) -> list:
    """
    Compute PageRank importance scores for all sentences in one pass.
    Called ONCE per document, not per batch.
    Uses spaCy pipe() for fast batch processing instead of one-by-one parsing.
    High score = sentence is structurally central = likely a key commitment.
    Returns list of floats (0-1 normalised) matching len(texts).
    """
    n = len(texts)
    if not HAS_SPACY or not HAS_NX or n == 0:
        return [0.0] * n
    try:
        # Use spaCy pipe() for batch processing — much faster than doc-by-doc
        truncated = [t[:300] for t in texts]
        dep_tokens = []
        for doc in _nlp.pipe(truncated, batch_size=64,
                              disable=["ner", "textcat"]):
            tokens = frozenset(
                t.lemma_.lower() for t in doc
                if t.dep_ in ("nsubj", "dobj", "ROOT") and len(t.lemma_) > 2
            )
            dep_tokens.append(tokens)

        # Build edges — only add edge if overlap is non-empty
        # Use set intersection which is fast
        G = _nx.DiGraph()
        G.add_nodes_from(range(n))
        for i in range(n):
            if not dep_tokens[i]:
                continue
            for j in range(n):
                if i != j and dep_tokens[i] & dep_tokens[j]:
                    G.add_edge(j, i)

        if G.number_of_edges() == 0:
            return [0.0] * n

        pr = _nx.pagerank(G, alpha=0.85, max_iter=100)
        scores = [pr.get(i, 0.0) for i in range(n)]
        mn, mx = min(scores), max(scores)
        if mx > mn:
            scores = [(s - mn) / (mx - mn) for s in scores]
        return [round(s, 4) for s in scores]
    except Exception as e:
        logging.warning(f"PageRank scoring failed: {e}")
        return [0.0] * n


# ── Classify via GPU classifier service ───────────────────────────────────────
def classify_with_gpu(
    chunks: List[Dict],
    policy_id: str,
    classifier_url: str,
) -> List[Dict]:
    """
    Send chunks to the GPU classifier API.
    Returns list of annotated chunks.
    """
    sentences = [c["text"] for c in chunks]
    try:
        resp = requests.post(
            f"{classifier_url}/kma/classify",
            json={"sentences": sentences, "policy_id": policy_id},
            timeout=120,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"    Classifier API failed: {e} — falling back to Mistral")
        return []

    results = data.get("results", [])

    # Compute PageRank once for all chunks
    pr_scores = _compute_pagerank_scores([c["text"] for c in chunks])

    annotated = []
    for i, (chunk, result) in enumerate(zip(chunks, results)):
        level = result.get("level", "Unsure")
        cls   = result.get("class", "Miscellaneous")

        # Override with hedge detection
        if _hedge_re.search(chunk["text"]):
            level = "Unsure"

        # Apply class correction rules
        cls = _correct_class(chunk["text"], cls)

        has_metric    = result.get("has_metric", bool(_metric_re.search(chunk["text"])))
        pr_score      = result.get("pagerank_score", pr_scores[i])

        annotated.append({
            **chunk,
            "level":           level,
            "class":           cls,
            "labels_str":      f"Level: {level}, Class: {cls}",
            "lv_confidence":   result.get("lv_confidence", 0.0),
            "cl_confidence":   result.get("cl_confidence", 0.0),
            "is_direct":       result.get("is_direct", True),
            "has_metric":      has_metric,
            "pagerank_score":  pr_score,
            "flagged_by":      "classifier",
            "needs_review":    result.get("needs_mistral_review", False),
        })
    return annotated


# ── Classify via Mistral (fallback / bootstrap) ────────────────────────────────
# PageRank disabled by default for speed — enable with --pagerank flag
_USE_PAGERANK = False

VALID_LEVELS  = {"Policy Action", "Policy Outcome", "Unsure"}
VALID_CLASSES = {
    "Area", "Emissions", "Site Status", "Spending", "Policy Action",
    "Knowledge Resource", "Practical Resource",
    "Environment Quality", "Miscellaneous",
}


def _normalise_item(item: dict) -> dict:
    """
    Normalise a single classification item.
    Handles uppercase keys (LEVEL/CLASS), prefixed values (Level: X),
    and invalid label names.
    """
    # case-insensitive key lookup
    def get(d, *keys):
        for k in keys:
            for dk in d:
                if dk.lower() == k.lower():
                    return d[dk]
        return None

    level = str(get(item, "level") or "Unsure").strip()
    cls   = str(get(item, "class") or "Miscellaneous").strip()

    # strip prefixes like "Level: Policy Action"
    level = re.sub(r'(?i)^level:\s*', '', level).strip()
    cls   = re.sub(r'(?i)^class:\s*', '', cls).strip()

    # map common mistral synonyms to valid labels
    LEVEL_MAP = {
        "policy action":  "Policy Action",
        "policy outcome": "Policy Outcome",
        "unsure":         "Unsure",
        "uncertain":      "Unsure",
        "unclear":        "Unsure",
        "ambiguous":      "Unsure",
    }
    CLASS_MAP = {
        "area":                 "Area",
        "emissions":            "Emissions",
        "site status":          "Site Status",
        "spending":             "Spending",
        "policy action":        "Policy Action",
        "knowledge resource":   "Knowledge Resource",
        "practical resource":   "Practical Resource",
        "environment quality":  "Environment Quality",
        "environmental quality":"Environment Quality",
        "miscellaneous":        "Miscellaneous",
        "misc":                 "Miscellaneous",
        "other":                "Miscellaneous",
        "resource/report":      "Knowledge Resource",
        "resource":             "Knowledge Resource",
        "report":               "Knowledge Resource",
    }

    level = LEVEL_MAP.get(level.lower(), "Unsure")
    cls   = CLASS_MAP.get(cls.lower(),   "Miscellaneous")

    is_direct = get(item, "is_direct")
    if is_direct is None:
        is_direct = True
    return {
        "level":     level,
        "class":     cls,
        "is_direct": bool(is_direct),
    }


def _correct_class(text: str, cls: str) -> str:
    """
    Rule-based class correction.
    Prevents topical words from overriding structural class signals.
    """
    text_lower = text.lower()

    # Override Spending — only keep if sentence has actual money/budget signal
    if cls == "Spending":
        spending_signals = [
            r"€\s*\d",                          # euro amount
            r"\$\s*\d",                          # dollar amount
            r"\d[\d,\.]*\s*(million|billion|thousand)\s*(euro|dollar|\$|€)",
            r"budget", r"funding\s+(of|worth|totalling)",
            r"per\s+annum", r"annual\s+payment",
            r"compensation\s+of", r"grant\s+of",
            r"financial\s+support\s+of",
        ]
        has_money = any(re.search(p, text, re.IGNORECASE) for p in spending_signals)
        if not has_money:
            # Resource mobilisation without figures → Policy Action class
            if any(s in text_lower for s in [
                "mobiliz", "mobilise", "financial resources", "funding sources",
                "resource mobilization", "financial and non-financial"
            ]):
                cls = "Policy Action"
            else:
                cls = "Miscellaneous"

    # Override Emissions when sentence is NOT a real emissions commitment
    # Real Emissions = specific target, reduction figure, or named obligation
    if cls == "Emissions":
        emissions_commitment_signals = [
            r"\b\d+\s*%",                    # percentage figure
            r"\b(reduce|reduction|cut|decrease|limit|cap)\b.{0,40}\b(emission|co2|carbon|greenhouse)",
            r"\b(emission|co2|carbon)\b.{0,40}\b(target|goal|limit|threshold|neutral)",
            r"\bcarbon\s+neutral",
            r"\bnet\s+zero",
            r"\b(80|95)\s*%\b",              # common EU climate targets
            r"\bby\s+20[3-9]\d\b",          # future deadline
        ]
        has_commitment = any(
            re.search(p, text_lower, re.IGNORECASE)
            for p in emissions_commitment_signals
        )
        if not has_commitment:
            cls = "Environment Quality"  # demote to general env quality

    # Legal/regulatory references → Policy Action class (not Area or Miscellaneous)
    legal_signals = [
        "legal consideration", "legal obligation", "eu and international",
        "national, eu and international", "habitats directive",
        "birds directive", "water framework directive",
        "environmental impact assessment", "article 6",
        "under the directive", "under eu law",
        "infringement", "court of justice",
    ]
    if cls in ("Area", "Miscellaneous") and any(s in text_lower for s in legal_signals):
        return "Policy Action"

    # If sentence is about producing guidance/plans/reports → Knowledge Resource
    knowledge_signals = [
        "introduce guidance", "provide guidance", "develop guidance",
        "produce guidance", "prepare guidance", "publish guidance",
        "introduce criteria", "develop criteria", "establish criteria",
        "management plan", "action plan", "research programme",
        "national database", "baseline survey", "assessment will be",
        "report will be", "review will be undertaken", "code of practice",
        "code of best practice", "framework will", "strategy will",
        "guidance on ", "guidance for ", "guidelines for",
    ]
    if any(s in text_lower for s in knowledge_signals):
        return "Knowledge Resource"

    # If sentence is about physical delivery to people → Practical Resource
    practical_signals = [
        "compensation package", "compensation will be",
        "relocation scheme", "will be relocated", "turf delivered",
        "alternative arrangement", "training will be", "training to",
        "will be provided with", "assistance will be",
    ]
    if any(s in text_lower for s in practical_signals):
        return "Practical Resource"

    # If sentence mentions WFD, water quality, habitat condition, biodiversity → Environment Quality
    env_signals = [
        "water framework directive", "wfd", "water quality",
        "environmental objective", "conservation objective",
        "favourable conservation", "ecological status",
        "habitat condition", "biodiversity target",
        "restoration of raised bog", "restoration of blanket bog",
        "ecosystem services", "biodiversity and", "loss of biodiversity",
        "natura 2000", "habitats directive", "birds directive",
        "carbon sequestration", "carbon stock", "greenhouse gas emissions",
        "peat oxidis", "rewetting", "carbon store",
    ]
    if any(s in text_lower for s in env_signals):
        return "Environment Quality"

    return cls


def _robust_parse_mistral(raw: str, batch_size: int) -> list:
    """
    Parse Zephyr/Mistral JSON response robustly.
    Handles: pretty-printed, compact, truncated, trailing commas,
    numbered dicts, uppercase keys, markdown fences.
    Four-step recovery:
      1. Direct json.loads
      2. Fix trailing commas + retry
      3. Truncate at last complete object + close array
      4. Extract complete objects with re.DOTALL (handles multiline)
    """
    if not raw:
        return []

    # Strip markdown fences
    raw = re.sub(r'```(?:json)?\s*', '', raw).strip()

    # Strip any text after the last ] — Zephyr sometimes adds "Explanation:" etc.
    last_bracket = raw.rfind(']')
    if last_bracket >= 0:
        raw = raw[:last_bracket + 1].strip()

    # Step 1: direct parse
    try:
        top = json.loads(raw)
        if isinstance(top, list):
            return [_normalise_item(i) for i in top if isinstance(i, dict)][:batch_size]
        if isinstance(top, dict):
            keys = list(top.keys())
            if all(str(k).isdigit() for k in keys):
                items = [v for v in top.values() if isinstance(v, dict)]
            elif any(str(k).lower() in ("level","class","is_direct") for k in keys):
                items = [top]
            else:
                items = [v for v in top.values() if isinstance(v, dict)]
            return [_normalise_item(i) for i in items][:batch_size]
    except Exception:
        pass

    # Step 2: fix trailing commas
    try:
        top = json.loads(re.sub(r',\s*([}\]])', r'\1', raw))
        if isinstance(top, list):
            return [_normalise_item(i) for i in top if isinstance(i, dict)][:batch_size]
    except Exception:
        pass

    # Step 3: close at last complete object
    last_close = raw.rfind('}')
    if last_close > 0:
        candidate = raw[:last_close + 1].rstrip().rstrip(',')
        if not candidate.lstrip().startswith('['):
            candidate = '[' + candidate
        if not candidate.rstrip().endswith(']'):
            candidate = candidate + ']'
        try:
            top = json.loads(re.sub(r',\s*([}\]])', r'\1', candidate))
            if isinstance(top, list) and top:
                return [_normalise_item(i) for i in top if isinstance(i, dict)][:batch_size]
        except Exception:
            pass

    # Step 4: extract complete objects with re.DOTALL (handles multiline pretty-print)
    complete = re.findall(
        r'\{[^{}]*?"level"\s*:\s*"[^"]*"[^{}]*?"class"\s*:\s*"[^"]*"[^{}]*?\}',
        raw, re.DOTALL
    )
    recovered = []
    for obj_str in complete:
        try:
            recovered.append(json.loads(obj_str))
        except Exception:
            try:
                recovered.append(json.loads(re.sub(r',\s*([}\]])', r'\1', obj_str)))
            except Exception:
                pass
    if recovered:
        return [_normalise_item(i) for i in recovered][:batch_size]

    return []


def classify_with_mistral(
    chunks: List[Dict],
    batch_size: int = 5,
) -> List[Dict]:
    """
    Classify sentences using local Zephyr via Ollama.
    Batch size kept small (4) to stay within context window.
    PageRank computed once upfront for all chunks, not per batch.
    """
    annotated = []
    total = len(chunks)

    # KPI scores always computed (fast regex)
    # PageRank disabled by default — too slow for large docs (opt-in via flag)
    all_texts     = [c["text"] for c in chunks]
    all_kpi_scores = [bool(_metric_re.search(t)) for t in all_texts]
    all_pr_scores  = _compute_pagerank_scores(all_texts) if _USE_PAGERANK else [0.0] * len(chunks)

    # ── Dynamic batching based on prompt length ──────────────────────────────
    # Build batches so total sentence text never exceeds MAX_PROMPT_CHARS.
    # This prevents Zephyr from truncating mid-JSON on long/non-ASCII sentences.
    MAX_PROMPT_CHARS = 300   # force 1-2 sentences per batch — avoids truncation

    def _safe_truncate(text: str, limit: int = 200) -> str:
        """Truncate at word boundary, halve limit for non-ASCII heavy text."""
        non_ascii = sum(1 for c in text if ord(c) > 127)
        effective_limit = limit // 2 if non_ascii / max(len(text), 1) > 0.08 else limit
        text = text[:effective_limit].replace(chr(10), " ").strip()
        if len(text) == effective_limit and " " in text:
            text = text[:text.rfind(" ")]
        return text

    batches = []
    current_batch, current_len = [], 0
    for chunk in chunks:
        slen = len(_safe_truncate(chunk["text"])) + 10  # +10 for "N. " prefix
        if current_batch and current_len + slen > MAX_PROMPT_CHARS:
            batches.append(current_batch)
            current_batch, current_len = [chunk], slen
        else:
            current_batch.append(chunk)
            current_len += slen
    if current_batch:
        batches.append(current_batch)

    total_batches = len(batches)
    print(f"    → {total} chunks in {total_batches} batches "
          f"(avg {total/total_batches:.1f} chunks/batch)", flush=True)

    # Track global index for pr/kpi score lookup
    global_idx = 0

    for batch_num, batch in enumerate(batches, 1):
        n = len(batch)

        # Progress line — updates in place
        pages_in_batch = sorted({c.get("page", "?") for c in batch})
        page_str = f"p{pages_in_batch[0]}" if len(pages_in_batch) == 1                    else f"p{pages_in_batch[0]}-{pages_in_batch[-1]}"
        print(
            f"\r    Batch {batch_num:>4}/{total_batches} "
            f"[{page_str:<8}] "
            f"{n} chunks ... ",
            end="", flush=True
        )

        sentences_text = "\n".join(
            f"{k+1}. {_safe_truncate(c['text'])}"
            for k, c in enumerate(batch)
        )

        ids = ", ".join(str(k+1) for k in range(n))
        prompt = f"""You are classifying sentences from environmental policy documents.
Return a JSON array with exactly {n} objects (ids {ids}). Nothing else.

LEVEL:
- "Policy Action"  = direct commitment with named actor + shall/will/must/requires
- "Policy Outcome" = result or goal the policy aims to achieve
- "Unsure"         = vague, hedged ("may", "could", "will be considered"), background, or description

CLASS:
- "Area"                = land area, hectares, geographic scope, afforestation targets
- "Emissions"           = specific CO2/GHG reduction target with % or deadline (NOT general climate text)
- "Site Status"         = named protected sites, SACs, NHAs, SPAs, designation or restoration of sites
- "Spending"            = actual euro/dollar amounts, annual payments, grants (NOT general resource mobilisation)
- "Policy Action"       = reference to another law, EU regulation, directive, or policy instrument
- "Knowledge Resource"  = plans, reports, databases, guidance documents, codes of practice, surveys produced
- "Practical Resource"  = compensation packages, relocation schemes, training delivery, physical actions to people
- "Environment Quality" = water quality, WFD, habitat condition, biodiversity, carbon sequestration, rewetting
- "Miscellaneous"       = does not clearly fit any above

is_direct: true = THIS policy commits to it. false = describes what another EU regulation/directive requires.

EXAMPLES (use these to calibrate):
Input: "The relevant authorities will introduce guidance and criteria for the identification and future management of peat areas."
Output: {{"id":1,"level":"Policy Action","class":"Knowledge Resource","is_direct":true}}

Input: "A comprehensive programme of restoration of raised bog SACs and NHAs shall be undertaken in partnership with landowners."
Output: {{"id":2,"level":"Policy Action","class":"Site Status","is_direct":true}}

Input: "A payment of €1,500 per annum for 15 years will be made to affected turf-cutters."
Output: {{"id":3,"level":"Policy Action","class":"Spending","is_direct":true}}

Input: "An aggregate reduction in CO2 emissions of at least 80% compared to 1990 levels by 2050."
Output: {{"id":4,"level":"Policy Outcome","class":"Emissions","is_direct":true}}

Input: "Forests can also help to provide temporary mitigation of climate change through sequestering carbon."
Output: {{"id":5,"level":"Unsure","class":"Environment Quality","is_direct":false}}

Input: "The aim is to ensure that Ireland meets its legal obligations under the Habitats Directive."
Output: {{"id":6,"level":"Unsure","class":"Policy Action","is_direct":false}}

Input: "Peat soils cover around 21% of the national land area."
Output: {{"id":7,"level":"Unsure","class":"Area","is_direct":false}}

Input: "The Cessation of Turf Cutting Compensation Scheme was established in April 2011."
Output: {{"id":8,"level":"Unsure","class":"Spending","is_direct":false}}

Input: "Peatland management influences the level, quantity and quality of water in the surrounding countryside."
Output: {{"id":9,"level":"Unsure","class":"Environment Quality","is_direct":false}}

Input: "DAFM will establish a national database of peatland ownership and turbary rights."
Output: {{"id":10,"level":"Policy Action","class":"Knowledge Resource","is_direct":true}}

Now classify:
{sentences_text}

Return ONLY the JSON array. Start with [ and end with ]"""

        payload = {
            "model":      "zephyr",
            "keep_alive": -1,
            "stream":     False,
            "messages": [
                {
                    "role":    "system",
                    "content": (
                        "You are a JSON-only policy annotator. "
                        "Output ONLY a valid JSON array. "
                        "No explanation. No markdown. No extra text. "
                        "Start with [ and end with ]."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            "options": {
                "num_ctx":     4096,
                "temperature": 0.0,
                "top_k":       1,
                "num_predict": 600,
            },
        }

        parsed = []
        try:
            res = requests.post(MISTRAL_URL, json=payload, timeout=180)
            res.raise_for_status()
            data = res.json()
            raw = (
                data.get("message", {}).get("content", "")
                or data.get("response", "")
            ).strip()
            parsed = _robust_parse_mistral(raw, len(batch))
            if not parsed:
                print(f"\r    Batch {batch_num:>4}/{total_batches} "
                      f"[{page_str:<8}] FAILED — no JSON recovered. "
                      f"Raw: {raw[:60]}", flush=True)
                logging.warning(
                    f"Batch {batch_num}/{total_batches}: "
                    f"Model returned no parseable JSON. Raw: {raw[:200]}"
                )
            elif len(parsed) < len(batch):
                print(f"\r    Batch {batch_num:>4}/{total_batches} "
                      f"[{page_str:<8}] PARTIAL — recovered {len(parsed)}/{len(batch)} chunks.",
                      flush=True)
        except requests.exceptions.Timeout:
            print(f"\r    Batch {batch_num:>4}/{total_batches} "
                  f"[{page_str:<8}] TIMEOUT — falling back to Unsure/Miscellaneous",
                  flush=True)
            logging.warning(f"Batch {batch_num}/{total_batches}: model timeout")
        except Exception as e:
            print(f"\r    Batch {batch_num:>4}/{total_batches} "
                  f"[{page_str:<8}] ERROR: {e}",
                  flush=True)
            logging.warning(f"Batch {batch_num}/{total_batches}: {e}")

        for j, chunk in enumerate(batch):
            r = parsed[j] if j < len(parsed) else {}
            level = r.get("level", "Unsure")
            cls   = r.get("class", "Miscellaneous")

            # override with hedge detection regardless of model output
            if _hedge_re.search(chunk["text"]):
                level = "Unsure"

            got_real_result = bool(parsed) and j < len(parsed)
            cls        = _correct_class(chunk["text"], cls)
            has_metric = all_kpi_scores[global_idx + j]
            pr_score   = all_pr_scores[global_idx + j]

            annotated.append({
                **chunk,
                "level":          level,
                "class":          cls,
                "labels_str":     f"Level: {level}, Class: {cls}",
                "lv_confidence":  0.75 if got_real_result else 0.5,
                "cl_confidence":  0.75 if got_real_result else 0.5,
                "is_direct":      r.get("is_direct", not bool(_indirect_re.search(chunk["text"]))),
                "has_metric":     has_metric,
                "pagerank_score": pr_score,
                "flagged_by":     "zephyr" if got_real_result else "zephyr_failed",
                "needs_review":   True,
            })

        global_idx += n

    # Final newline after progress line
    print(f"\r    Done — {len(annotated)} chunks classified.           ", flush=True)
    return annotated


# ── Excel builder ─────────────────────────────────────────────────────────────
def _fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _hdr(ws, headers, colours, row=1):
    for col, (h, bg) in enumerate(zip(headers, colours), 1):
        c = ws.cell(row, col, h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center",
                                vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 28


def build_excel(
    policy_id, pdf_name, total_pages, skipped_pages,
    annotated, unlabelled, output_path, metadata,
    classifier_used,
):
    wb = Workbook()
    total = len(annotated) + len(unlabelled)
    cov = round(len(annotated) / total * 100, 1) if total else 0

    # ── Summary ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"KMA Annotation Report — Policy {policy_id}"
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

    meta_rows = [
        ("Policy ID",           policy_id),
        ("PDF File",            pdf_name),
        ("Country",             metadata.get("country") or "—"),
        ("Language",            metadata.get("language") or "—"),
        ("Governance Level",    metadata.get("governance_level") or "—"),
        ("Total Pages",         total_pages),
        ("Skipped Pages",       skipped_pages),
        ("Total Chunks",        total),
        ("Annotated Chunks",    len(annotated)),
        ("Unlabelled Chunks",   len(unlabelled)),
        ("Coverage",            f"{cov}%"),
        ("Classifier Used",     classifier_used),
    ]
    for i, (k, v) in enumerate(meta_rows, 3):
        ws.cell(i, 1, k).font = Font(bold=True, name="Arial", size=10)
        ws.cell(i, 2, v).font = Font(name="Arial", size=10)

    HDR = 17
    ws.row_dimensions[HDR - 1].height = 10
    _hdr(ws,
         ["Label", "Category", "Count", "% of Annotated",
          "% of Total", "Colour"],
         ["2E4057"] * 6, row=HDR)

    lc = defaultdict(int)
    for c in annotated:
        lc[c.get("labels_str", "")] += 1

    for off, (label, colour) in enumerate(LABEL_COLOURS.items()):
        r = HDR + 1 + off
        count = lc.get(label, 0)
        cat = label.split(":")[0].strip()
        pct_a = round(count / len(annotated) * 100, 1) if annotated else 0
        pct_t = round(count / total * 100, 1) if total else 0
        for col, v in enumerate(
            [label, cat, count, f"{pct_a}%", f"{pct_t}%", ""], 1
        ):
            cell = ws.cell(r, col, v)
            cell.font = Font(name="Arial", size=10)
            cell.border = _border()
            cell.fill = _fill(colour) if col in (1, 6) else _fill("FAFAFA")
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4, 5) else "left"
            )

    for col, w in zip("ABCDEF", [28, 16, 10, 16, 14, 12]):
        ws.column_dimensions[col].width = w

    # ── Annotations ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Annotations")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    _hdr(ws2,
         ["Chunk #", "Page", "Level", "Class",
          "lv_conf", "cl_conf", "Direct?", "KPI?", "PageRank",
          "Flagged by", "Text", "Chars",
          "Correct? (Y/N/Partial)", "Notes"],
         ["1F4E79"] * 12 + ["2E7D32", "2E7D32"])

    for row_off, chunk in enumerate(annotated, 2):
        label_key = chunk.get("labels_str", "")
        colour = LABEL_COLOURS.get(label_key, "FFFFFF")
        pr = chunk.get("pagerank_score", 0.0)
        vals = [
            chunk["chunk_index"],
            chunk.get("page", ""),
            chunk.get("level", ""),
            chunk.get("class", ""),
            chunk.get("lv_confidence", ""),
            chunk.get("cl_confidence", ""),
            "Y" if chunk.get("is_direct", True) else "N",
            "Y" if chunk.get("has_metric", False) else "",
            f"{pr:.3f}" if pr else "",
            chunk.get("flagged_by", ""),
            chunk["text"],
            len(chunk["text"]),
            "Y" if not chunk.get("needs_review") else "",
            "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row_off, col, v)
            cell.font = Font(name="Arial", size=9)
            cell.border = _border()
            if col == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col in (11, 12):
                cell.fill = _fill("FFFFFF")
            else:
                cell.fill = _fill(colour)
                cell.alignment = Alignment(
                    horizontal="center"
                    if col in (1, 2, 5, 6, 10) else "left"
                )
        ws2.row_dimensions[row_off].height = max(
            15, min(80, len(chunk["text"]) // 8)
        )

    for col, w in zip("ABCDEFGHIJKLMN",
                      [9, 7, 18, 18, 8, 8, 8, 7, 9, 12, 65, 7, 22, 28]):
        ws2.column_dimensions[col].width = w

    # ── Unlabelled ──────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Unlabelled")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"
    _hdr(ws3,
         ["Chunk #", "Page", "Text", "Chars",
          "Suggested Level", "Suggested Class", "Notes"],
         ["616161"] * 4 + ["2E7D32"] * 3)

    for row_off, chunk in enumerate(unlabelled, 2):
        vals = [
            chunk["chunk_index"], chunk.get("page", ""),
            chunk["text"], len(chunk["text"]), "", "", "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row_off, col, v)
            cell.font = Font(name="Arial", size=9)
            cell.border = _border()
            if col == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = _fill("F8F8F8")
            elif col in (5, 6, 7):
                cell.fill = _fill("FFFFFF")
            else:
                cell.fill = _fill("F8F8F8")
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 4) else "left"
                )
        ws3.row_dimensions[row_off].height = max(
            15, min(80, len(chunk["text"]) // 8)
        )

    for col, w in zip("ABCDEFG", [9, 7, 70, 8, 22, 22, 28]):
        ws3.column_dimensions[col].width = w

    wb.save(str(output_path))


# ── Per-policy processor ───────────────────────────────────────────────────────
def process_policy(
    policy_dir: Path,
    metadata: Dict,
    chunk_size: int,
    chunk_overlap: int,
    classifier_url: str,
    mistral_only: bool,
) -> bool:
    policy_id = policy_dir.name
    pdf_files = [
        f for f in policy_dir.iterdir()
        if f.is_file() and f.suffix.lower() == ".pdf"
    ]
    if not pdf_files:
        print(f"  [{policy_id}] No PDF — skipping")
        return False

    pdf_path = pdf_files[0]
    excel_path = policy_dir / f"{pdf_path.stem}_annotations.xlsx"

    if excel_path.exists():
        print(f"  [{policy_id}] Already annotated — skipping")
        return True

    print(f"  [{policy_id}] Extracting {pdf_path.name}", end="", flush=True)

    pages = extract_text_from_pdf(pdf_path)
    total_pages  = len(pages)
    skipped_pages = sum(1 for p in pages if p["source"] == "skipped_graphic")
    print(f" — {total_pages}p ({skipped_pages} skipped)", end="", flush=True)

    chunks = chunk_pages(pages, chunk_size, chunk_overlap)
    if not chunks:
        print(f"\n  [{policy_id}] No text extracted")
        return False

    to_classify, skipped = pre_filter(chunks)
    print(f", {len(chunks)} chunks ({len(to_classify)} to classify)", flush=True)

    # Check if GPU classifier is available
    classifier_available = False
    if not mistral_only:
        try:
            r = requests.get(f"{classifier_url}/kma/health", timeout=5)
            if r.ok and r.json().get("status") == "ready":
                classifier_available = True
        except Exception:
            pass

    if classifier_available:
        print(f"  [{policy_id}] Using GPU classifier...", end="", flush=True)
        annotated = classify_with_gpu(to_classify, policy_id, classifier_url)
        classifier_used = "GPU classifier (step3)"
        if not annotated:
            # fallback
            annotated = classify_with_mistral(to_classify)
            classifier_used = "Zephyr (GPU fallback)"
    else:
        print(f"  [{policy_id}] Using Zephyr (bootstrap mode)...",
              end="", flush=True)
        annotated = classify_with_mistral(to_classify)
        classifier_used = "Zephyr (Phase 1 bootstrap)"

    # Separate direct/annotated from indirect/unlabelled
    direct    = [c for c in annotated if c.get("is_direct", True)]
    indirect  = [c for c in annotated if not c.get("is_direct", True)]
    unlabelled = skipped + indirect

    cov = round(len(direct) / len(chunks) * 100, 1) if chunks else 0
    print(f" done. {len(direct)} direct, {len(indirect)} indirect, "
          f"{cov}% coverage")

    build_excel(
        policy_id=policy_id,
        pdf_name=pdf_path.name,
        total_pages=total_pages,
        skipped_pages=skipped_pages,
        annotated=direct,
        unlabelled=unlabelled,
        output_path=excel_path,
        metadata=metadata,
        classifier_used=classifier_used,
    )
    print(f"  [{policy_id}] Saved: {excel_path.name}")
    return True


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",    required=True)
    parser.add_argument("--limit",    type=int, default=10)
    parser.add_argument("--country",  default=None)
    parser.add_argument("--language", default=None)
    parser.add_argument("--governance-level", default=None)
    parser.add_argument("--chunk-size",    type=int, default=350)
    parser.add_argument("--chunk-overlap", type=int, default=50)
    parser.add_argument("--classifier-url", default=CLASSIFIER_URL)
    parser.add_argument("--mistral-only",   action="store_true")
    parser.add_argument("--pagerank",       action="store_true",
                        help="Compute PageRank scores (slower, adds ~1min per 500 chunks)")
    args = parser.parse_args()

    base = Path(args.input)
    if not base.exists():
        print(f"Folder not found: {base}")
        sys.exit(1)

    policy_dirs = sorted(
        [d for d in base.iterdir() if d.is_dir()],
        key=lambda d: int(d.name) if d.name.isdigit() else d.name,
    )
    selected = policy_dirs[:args.limit] if args.limit > 0 else policy_dirs
    metadata = {
        "country": args.country,
        "language": args.language,
        "governance_level": args.governance_level,
    }

    # Set PageRank flag globally
    global _USE_PAGERANK
    _USE_PAGERANK = args.pagerank

    # Check classifier
    if not args.mistral_only:
        try:
            r = requests.get(f"{args.classifier_url}/kma/health", timeout=5)
            if r.ok and r.json().get("status") == "ready":
                print(f"GPU classifier: READY at {args.classifier_url}")
            else:
                print(f"GPU classifier: NOT READY — using Mistral fallback")
        except Exception:
            print(f"GPU classifier: UNREACHABLE at {args.classifier_url}")
            print(f"               — using Mistral fallback (Phase 1 mode)")

    print(f"\nProcessing {len(selected)} policies...\n")
    ok, fail = 0, 0
    for d in selected:
        try:
            if process_policy(
                d, metadata,
                args.chunk_size, args.chunk_overlap,
                args.classifier_url, args.mistral_only,
            ):
                ok += 1
            else:
                fail += 1
        except Exception as e:
            print(f"  [{d.name}] ERROR: {e}")
            traceback.print_exc()
            fail += 1

    print(f"\nDone. {ok} succeeded, {fail} failed.")


if __name__ == "__main__":
    main()
