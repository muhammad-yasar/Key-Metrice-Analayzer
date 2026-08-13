"""
step1_prepare_training_data.py
==============================
Reads reviewed annotation Excel files and builds training_data.csv.
Handles both:
  - Flat folder: ~/files/annotated_excel_file/*.xlsx
  - Nested folder: ~/files/exported_files/*/policy_id/*_annotations.xlsx

Run:
    python step1_prepare_training_data.py --input ~/files/annotated_excel_file
    python step1_prepare_training_data.py --input ~/files/exported_files

Output:
    training_data.csv
    training_data_stats.json
"""

import pandas as pd
import glob
import json
import argparse
import os
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

# ── Label taxonomy (must match step4) ────────────────────────────────────────
LEVEL_LABELS = ["Policy Action", "Policy Outcome", "Unsure"]

CLASS_LABELS = [
    "Area", "Emissions", "Site Status", "Spending", "Policy Action",
    "Knowledge Resource", "Practical Resource", "Environment Quality",
    "Miscellaneous",
]

# Minimum examples per class to flag as low — training will proceed anyway
MIN_WARN = 15



# ── Note parser — extracts label corrections from free-text reviewer notes ────
_LABEL_ALIASES = {
    "environmental quality": "Environment Quality",
    "env quality": "Environment Quality",
    "knowledge resources": "Knowledge Resource",
    "practical resources": "Practical Resource",
    "outcome": "Policy Outcome",
    "action": "Policy Action",
}
_ALL_LABELS_SORTED = sorted(
    [(l, "level") for l in LEVEL_LABELS] +
    [(l, "class") for l in CLASS_LABELS] +
    [(_a, "both") for _a in _LABEL_ALIASES],
    key=lambda x: -len(x[0])
)

def _parse_note(note: str):
    """
    Extract corrected level and/or class from free-text reviewer notes.
    Handles: 'Policy action rather than outcome',
             'Environmental quality rather than miscellaneous',
             'Outcome rather than Action', 'Should be Spending', etc.
    """
    import re as _re
    note_clean = _re.sub(r'["\']', '', note.strip())
    note_lower = note_clean.lower()

    def resolve(label):
        label_l = label.lower()
        if label_l in _LABEL_ALIASES:
            resolved = _LABEL_ALIASES[label_l]
            if resolved in LEVEL_LABELS: return resolved, "level"
            return resolved, "class"
        if label in LEVEL_LABELS: return label, "level"
        if label in CLASS_LABELS: return label, "class"
        return label, "both"

    mentions, used_spans = [], []
    for label, _ in _ALL_LABELS_SORTED:
        for m in _re.finditer(_re.escape(label.lower()), note_lower):
            if not any(s <= m.start() < e or s < m.end() <= e for s,e in used_spans):
                rl, lt = resolve(label)
                mentions.append((m.start(), m.end(), rl, lt))
                used_spans.append((m.start(), m.end()))
    mentions.sort()

    neg_starts = [m.start() for p in [r'\brather\s+than\b',r'\bnot\b',r'\binstead\s+of\b']
                  for m in _re.finditer(p, note_lower)]
    pos_starts = [m.start() for p in [r'\bbetter\s+as\b',r'\bshould\s+be\b',r'\bactually\b']
                  for m in _re.finditer(p, note_lower)]

    is_wrong    = lambda pos: any(ns < pos and pos - ns < 35 for ns in neg_starts)
    is_explicit = lambda pos: any(ps < pos and pos - ps < 35 for ps in pos_starts)

    def forced_type(end):
        after = note_lower[end:end+15]
        if _re.match(r'\s*class\b', after): return "class"
        if _re.match(r'\s*level\b', after): return "level"
        return None

    explicit = [(p,e,l,t) for p,e,l,t in mentions if is_explicit(p) and not is_wrong(p)]
    plain    = [(p,e,l,t) for p,e,l,t in mentions if not is_wrong(p) and not is_explicit(p)]
    candidates = explicit if explicit else plain

    found_level = found_class = None
    for p, e, label, ltype in candidates:
        ft = forced_type(e)
        if ft == "class" and not found_class:   found_class = label; continue
        if ft == "level" and not found_level:   found_level = label; continue
        if ltype == "level" and not found_level: found_level = label
        elif ltype == "class" and not found_class: found_class = label
        elif ltype == "both":
            if "class" in note_lower[:p+20] and not found_class: found_class = label
            elif "level" in note_lower[:p+20] and not found_level: found_level = label
            elif label in LEVEL_LABELS and not found_level: found_level = label
            elif not found_class: found_class = label

    if found_level == "Policy Action" and found_class == "Policy Action":
        if "class" in note_lower: found_level = None
        else: found_class = None

    return found_level, found_class


def read_annotations_xlsx(path: Path) -> list:
    """
    Read an annotations Excel file and return list of valid training rows.
    Handles both Labels column format and separate Level/Class columns.
    """
    rows = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  SKIP {path.name}: cannot open — {e}")
        return rows

    if "Annotations" not in wb.sheetnames:
        print(f"  SKIP {path.name}: no Annotations sheet")
        return rows

    ws = wb["Annotations"]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return rows

    headers = [str(h or "").strip() for h in all_rows[0]]

    # Find column indices
    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    correct_idx = col("Correct? (Y/N/Partial)")
    text_idx    = col("Text")
    level_idx   = col("Level")
    class_idx   = col("Class")
    labels_idx  = col("Labels")        # old format
    notes_idx   = col("Notes")

    if correct_idx is None or text_idx is None:
        print(f"  SKIP {path.name}: missing required columns")
        return rows

    kept = skipped_n = skipped_blank = skipped_invalid = 0

    for row in all_rows[1:]:
        if len(row) <= correct_idx:
            continue

        verdict = str(row[correct_idx] or "").strip()
        if verdict == "N":
            skipped_n += 1
            continue
        if verdict not in ("Y", "Partial"):
            skipped_blank += 1
            continue

        if text_idx >= len(row):
            skipped_blank += 1
            continue
        text = str(row[text_idx] or "").strip()
        if not text or text == "nan":
            skipped_blank += 1
            continue

        # Get level and class
        level, cls = None, None

        # Method 1: separate Level and Class columns (current format)
        if level_idx is not None and class_idx is not None:
            if level_idx < len(row):
                level = str(row[level_idx] or "").strip()
            if class_idx < len(row):
                cls = str(row[class_idx] or "").strip()

        # Method 2: combined Labels column (old format)
        elif labels_idx is not None and labels_idx < len(row):
            label_str = str(row[labels_idx] or "").strip()
            for part in label_str.split(","):
                part = part.strip()
                if part.startswith("Level:"):
                    level = part.replace("Level:", "").strip()
                elif part.startswith("Class:"):
                    cls = part.replace("Class:", "").strip()

        # Check for manual correction in Notes column
        # Handles free-text notes like "Policy action rather than outcome"
        # "Environmental quality rather than miscellaneous", "Should be Spending" etc.
        if notes_idx is not None and notes_idx < len(row):
            note = str(row[notes_idx] or "").strip()
            if note and note != "nan":
                corrected_level, corrected_class = _parse_note(note)
                if corrected_level:
                    level = corrected_level
                if corrected_class:
                    cls = corrected_class

        # Validate
        if level not in LEVEL_LABELS:
            skipped_invalid += 1
            continue
        if cls not in CLASS_LABELS:
            skipped_invalid += 1
            continue

        rows.append({
            "text":        text,
            "level_label": level,
            "class_label": cls,
            "source_file": path.name,
        })
        kept += 1

    print(f"  {path.name:<55} kept={kept:<4} "
          f"N={skipped_n:<4} blank={skipped_blank:<4} bad_label={skipped_invalid}")
    return rows


def find_xlsx_files(input_dir: str) -> list:
    """Find all annotation Excel files — flat or nested folder."""
    base = Path(input_dir)
    files = []

    # Flat folder: *.xlsx directly in input_dir
    flat = list(base.glob("*.xlsx"))
    if flat:
        files.extend(flat)

    # Nested: *_annotations.xlsx in subdirectories
    nested = list(base.glob("**/*_annotations.xlsx"))
    nested = [f for f in nested if f not in files]
    files.extend(nested)

    return sorted(set(files))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  required=True,
                        help="Folder containing reviewed annotation Excel files")
    parser.add_argument("--output", default="training_data.csv")
    parser.add_argument("--append", action="store_true",
                        help="Append to existing CSV instead of overwriting")
    args = parser.parse_args()

    files = find_xlsx_files(args.input)
    if not files:
        print(f"No Excel files found in: {args.input}")
        return

    print(f"Found {len(files)} Excel files\n")
    print(f"{'File':<55} {'kept':<6} {'N':<6} {'blank':<7} bad_label")
    print("-" * 90)

    all_rows = []
    for f in files:
        all_rows.extend(read_annotations_xlsx(f))

    if not all_rows:
        print("\nNo valid training rows found.")
        print("Check that Correct? column has Y or Partial values.")
        return

    df_new = pd.DataFrame(all_rows)

    # Append or overwrite
    if args.append and os.path.exists(args.output):
        df_existing = pd.read_csv(args.output)
        df_out = pd.concat([df_existing, df_new], ignore_index=True)
        df_out = df_out.drop_duplicates(subset=["text", "level_label", "class_label"])
        print(f"\nAppended: {len(df_existing)} existing + {len(df_new)} new "
              f"= {len(df_out)} total (deduped)")
    else:
        df_out = df_new
        print(f"\nCreated: {len(df_out)} rows")

    df_out.to_csv(args.output, index=False)

    # Distribution report
    print("\n=== Level distribution ===")
    level_counts = df_out["level_label"].value_counts().to_dict()
    for label in LEVEL_LABELS:
        count = level_counts.get(label, 0)
        bar   = "█" * (count // 10)
        flag  = "  ← LOW" if count < MIN_WARN else ""
        print(f"  {label:<22} {count:>4}  {bar}{flag}")

    print("\n=== Class distribution ===")
    class_counts = df_out["class_label"].value_counts().to_dict()
    for label in CLASS_LABELS:
        count = class_counts.get(label, 0)
        bar   = "█" * (count // 10)
        flag  = "  ← LOW" if count < MIN_WARN else ""
        print(f"  {label:<24} {count:>4}  {bar}{flag}")

    # Readiness check
    low_levels  = [l for l in LEVEL_LABELS  if level_counts.get(l, 0) < MIN_WARN]
    low_classes = [c for c in CLASS_LABELS  if class_counts.get(c, 0) < MIN_WARN]
    blocked     = [c for c in CLASS_LABELS  if class_counts.get(c, 0) < 3]

    print()
    if blocked:
        print(f"BLOCKED — fewer than 3 examples (cannot train): {', '.join(blocked)}")
        print("Add more annotated policies or remove these classes from CLASS_LABELS.")
    elif low_levels or low_classes:
        print("PROCEED WITH CAUTION — some classes are low:")
        for l in low_levels:
            print(f"  Level: {l} ({level_counts.get(l,0)} examples)")
        for c in low_classes:
            print(f"  Class: {c} ({class_counts.get(c,0)} examples)")
        print("\nClass weighting in step2 will compensate. Proceed to step2.")
    else:
        print("READY — all labels have sufficient examples. Proceed to step2.")

    # Save stats
    stats = {
        "total_rows":         len(df_out),
        "files_processed":    len(files),
        "level_counts":       level_counts,
        "class_counts":       class_counts,
        "low_classes":        low_classes,
        "blocked":            blocked,
    }
    stats_path = args.output.replace(".csv", "_stats.json")
    with open(stats_path, "w") as f:
        json.dump(stats, f, indent=2)

    print(f"\nSaved: {args.output}")
    print(f"Saved: {stats_path}")
    print(f"\nNext step:")
    print(f"  python step2_train_classifier.py --data {args.output}")


if __name__ == "__main__":
    main()
