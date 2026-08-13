"""
check_distribution.py
=====================
Quick script to check class distribution of reviewed annotation files
before running step2 training.

Run:
    python scripts/check_distribution.py --input ./annotated_excel_file
"""

import argparse
from pathlib import Path
from openpyxl import load_workbook
from collections import Counter

LEVEL_LABELS = ["Policy Action", "Policy Outcome", "Unsure"]
CLASS_LABELS = [
    "Area", "Emissions", "Site Status", "Spending", "Policy Action",
    "Knowledge Resource", "Practical Resource", "Environment Quality",
    "Miscellaneous",
]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True,
                        help="Folder containing reviewed annotation Excel files")
    parser.add_argument("--min", type=int, default=30,
                        help="Minimum examples per class (default: 30)")
    args = parser.parse_args()

    base = Path(args.input)
    level_counts = Counter()
    class_counts  = Counter()
    total = 0

    for xlsx in sorted(base.glob("*.xlsx")):
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        if "Annotations" not in wb.sheetnames:
            print(f"  SKIP {xlsx.name} — no Annotations sheet")
            continue
        ws = wb["Annotations"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        try:
            correct_idx = headers.index("Correct? (Y/N/Partial)")
            level_idx   = headers.index("Level")
            class_idx   = headers.index("Class")
        except ValueError:
            print(f"  SKIP {xlsx.name} — missing columns")
            continue
        file_total = 0
        for row in rows[1:]:
            if len(row) <= max(correct_idx, level_idx, class_idx):
                continue
            verdict = str(row[correct_idx] or "").strip()
            if verdict in ("Y", "Partial"):
                level_counts[row[level_idx]] += 1
                class_counts[row[class_idx]]  += 1
                total += 1
                file_total += 1
        print(f"  {xlsx.name}: {file_total} usable rows")

    print(f"\nTotal usable rows: {total}")

    print(f"\nLevel distribution:")
    for k in LEVEL_LABELS:
        v    = level_counts.get(k, 0)
        bar  = "█" * (v // 5)
        flag = "  ← LOW" if v < args.min else ""
        print(f"  {k:<22} {v:>4}  {bar}{flag}")

    print(f"\nClass distribution:")
    for k in CLASS_LABELS:
        v    = class_counts.get(k, 0)
        bar  = "█" * (v // 5)
        flag = "  ← LOW" if v < args.min else ""
        print(f"  {k:<24} {v:>4}  {bar}{flag}")

    low      = [k for k in CLASS_LABELS if class_counts.get(k, 0) < args.min]
    blocked  = [k for k in CLASS_LABELS if class_counts.get(k, 0) < 5]

    print(f"\nDecision:")
    if blocked:
        print(f"  BLOCKED — fewer than 5 examples: {', '.join(blocked)}")
    elif low:
        print(f"  PROCEED WITH CAUTION — low classes: {', '.join(low)}")
        print(f"  Proceed to step1 → step2 (class weighting will compensate).")
    else:
        print(f"  READY — all classes have {args.min}+ examples. Proceed to step1.")


if __name__ == "__main__":
    main()
